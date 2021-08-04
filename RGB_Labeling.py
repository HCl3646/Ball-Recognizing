import math
import sys
import cv2
import numpy as np
import Definitions as df
import erasing_background as eb
from openpyxl import Workbook

mod = sys.modules[__name__]
video = cv2.VideoCapture("Golf_Test_1.mp4")

if not video.isOpened():
    sys.exit()

delay = df.getDelay(video)
idx, score, score_0 = np.zeros(3)
balls = list()
score_list = list()

while True:
    ret, img = video.read()

    if not ret:
        break

    img = cv2.GaussianBlur(img, (0, 0), 1.0)
    dst = cv2.inRange(img, np.array([90, 90, 90]), np.array([255, 255, 255]))
    src = cv2.cvtColor(dst, cv2.COLOR_GRAY2BGR)
    _, labels, stats, centroids = cv2.connectedComponentsWithStats(dst)
    pts = list()

    for i in range(len(stats)):
        x, y, w, h, area = stats[i]
        mx, my = centroids[i]
        if (h, w) < dst.shape and 500 * 4 < w * h < 3000 * 4 and 2 / 3 < w / h < 4 / 3 and 0.8 < area / (
                0.25 * math.pi * w * h) < 1.2:
            cv2.rectangle(src, (x, y, w, h), (0, 0, 255), 1)
            cv2.line(src, (int(mx), int(my)), (int(mx), int(my)), (0, 0, 255), 3)
            point = (mx, my, idx)
            pts.append(point)

                        
    for p in pts:
        cnt = 0
        length_0 = 9999999999
        temp = list()
        moving = True
        for j in balls:
            length = df.distance(j[ - 1][0], j[- 1][1], p[0], p[1])
            if length > 200:
                cnt += 1
            elif length < 5:
                moving = False
                break
            else:
                if length <= length_0:
                    length_0 = length
                    temp = j
                else:
                    continue
        if cnt == len(balls):
            balls.append([p])
        elif moving is True and cnt != len(balls):
            temp.append(p)
        else:
            continue

    for i in balls:
        if len(i) >= 2 and idx == i[-1][2]:
            if len(i) >= 3 and df.distance(i[-1][0], i[-1][1], i[-2][0], i[-2][1]) >= 12:
                g0 = df.gradient(i[- 2][0], i[- 2][1], i[- 3][0], i[- 3][1])
                g1 = df.gradient(i[- 1][0], i[- 1][1], i[- 2][0], i[- 2][1])
                if abs(math.atan(g0)-math.atan(g1)) >= math.pi/12:
                    score += 1
                    break
                    
            elif i[-1][2] - i[-2][2] > delay * 7:
                score += 1
                break

    for i in balls:
        for j in range(len(i)-1):
            cv2.line(src, (int(i[j][0]),int(i[j][1])),(int(i[j+1][0]),int(i[j+1][1])),(0,255,255),1)
    
    cv2.imshow('dst', src)
    
    idx += 1
    if score != score_0:
        print(int(score), int(idx))
        score_0 = score
        score_list.append(idx)
        cv2.imwrite("images/img_{}.jpg".format(int(idx)),src)

    if cv2.waitKey(delay) == 27:
        break

video.release()
cv2.destroyAllWindows()

write_wb = Workbook()
write_ws = write_wb.active
write_ws.append(score_list)
for i in balls:
    for j in i:
        (write_ws.cell(column=4*balls.index(i)+1, row= i.index(j)+3)).value = j[0]
        (write_ws.cell(column=4*balls.index(i)+2, row= i.index(j)+3)).value = j[1]
        (write_ws.cell(column=4*balls.index(i)+3, row= i.index(j)+3)).value = j[2]
write_wb.save('images/balls.xlsx')
#=(F2-F1)/(E2-E1)